"""
数据导出服务 — 交易记录/自选股/投资组合导出为 Excel/CSV
降级链: openpyxl → csv (标准库)
"""
import csv
import io
import logging
from typing import Any, List, Optional
from src.utils import now_et

logger = logging.getLogger(__name__)

# ── 可选依赖: openpyxl ──────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import (
        Alignment, Border, Font, PatternFill, Side,
    )
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    logger.info("openpyxl 未安装，导出将降级为 CSV")

# ── 样式常量 ─────────────────────────────────────────────────
if HAS_OPENPYXL:
    _HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    _HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    _HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
    _THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    _GREEN_FONT = Font(color="006100")
    _GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    _RED_FONT = Font(color="9C0006")
    _RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    _NUM_FMT_USD = '#,##0.00'
    _NUM_FMT_CNY = '¥#,##0.00'
    _NUM_FMT_PCT = '0.00%'


def _auto_width(ws) -> None:
    """自动调整列宽 (openpyxl)"""
    for col in ws.columns:
        max_len = 0
        col_letter = None
        for cell in col:
            if col_letter is None and hasattr(cell, 'column_letter'):
                col_letter = cell.column_letter
            try:
                cell_len = len(str(cell.value or ""))
                if cell_len > max_len:
                    max_len = cell_len
            except Exception as e:
                logger.debug("Silenced exception", exc_info=True)
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 10), 40)


def _style_header(ws, headers: List[str]) -> None:
    """统一表头样式"""
    ws.append(headers)
    for cell in ws[ws.max_row]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER


def _style_pnl_cell(cell, value: float) -> None:
    """PnL 单元格颜色编码: 正值绿色，负值红色"""
    cell.border = _THIN_BORDER
    if value > 0:
        cell.font = _GREEN_FONT
        cell.fill = _GREEN_FILL
    elif value < 0:
        cell.font = _RED_FONT
        cell.fill = _RED_FILL


def _write_csv(headers: List[str], rows: List[List[Any]]) -> io.BytesIO:
    """CSV 降级输出"""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    writer.writerows(rows)
    result = io.BytesIO(buf.getvalue().encode("utf-8-sig"))  # BOM for Excel
    result.seek(0)
    return result


# ── 公共 API ─────────────────────────────────────────────────


def export_trades(trades: List[dict], format: str = "xlsx") -> io.BytesIO:
    """导出交易记录

    Parameters
    ----------
    trades : list[dict]
        每条交易包含: symbol, action, quantity, price, total, executed_at, profit(可选)
    format : str
        "xlsx" (默认) 或 "csv"

    Returns
    -------
    io.BytesIO
        可直接发送给 Telegram 的文件字节流
    """
    headers = ["日期", "代码", "方向", "数量", "价格", "总额", "盈亏"]

    rows = []
    for t in trades:
        row = [
            t.get("executed_at", "")[:19],
            t.get("symbol", ""),
            t.get("action", ""),
            t.get("quantity", 0),
            t.get("price", 0),
            t.get("total", 0),
            t.get("profit", 0),
        ]
        rows.append(row)

    if format == "csv" or not HAS_OPENPYXL:
        return _write_csv(headers, rows)

    # Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "交易记录"

    # 标题行
    ws.append([f"交易记录导出 — {now_et().strftime('%Y-%m-%d %H:%M')}"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws[1][0].font = Font(name="Arial", bold=True, size=14)
    ws.append([])  # 空行

    _style_header(ws, headers)

    for row in rows:
        ws.append(row)
        row_idx = ws.max_row
        # 价格/总额格式
        ws.cell(row=row_idx, column=5).number_format = _NUM_FMT_USD
        ws.cell(row=row_idx, column=6).number_format = _NUM_FMT_USD
        # 盈亏颜色
        pnl_cell = ws.cell(row=row_idx, column=7)
        pnl_cell.number_format = _NUM_FMT_USD
        _style_pnl_cell(pnl_cell, row[6] or 0)
        # 方向颜色
        action_cell = ws.cell(row=row_idx, column=3)
        if row[2] == "BUY":
            action_cell.font = Font(color="006100", bold=True)
        elif row[2] == "SELL":
            action_cell.font = Font(color="9C0006", bold=True)
        # 通用边框
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col_idx).border = _THIN_BORDER

    _auto_width(ws)
    ws.freeze_panes = "A4"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── 记账数据导出 ─────────────────────────────────────────────

def export_expenses(
    expenses: List[dict],
    summary: Optional[dict] = None,
    format: str = "xlsx",
) -> io.BytesIO:
    """导出记账数据 — 支出+收入明细表 + 月度汇总sheet

    Parameters
    ----------
    expenses : list[dict]
        每条记录: date, type(income/expense), category, amount, note
    summary : dict, optional
        月度汇总数据 — 包含 months 列表:
        [{month, total_expense, total_income, net, budget, budget_pct}]
    format : str
        "xlsx" (默认) 或 "csv"
    """
    headers = ["日期", "类型", "分类", "金额", "备注"]

    rows = []
    for e in expenses:
        row_type = "收入" if e.get("type") == "income" else "支出"
        rows.append([
            e.get("date", ""),
            row_type,
            e.get("category", "其他"),
            e.get("amount", 0),
            e.get("note", ""),
        ])

    if format == "csv" or not HAS_OPENPYXL:
        # CSV 模式: 明细 + 汇总追加到末尾
        all_rows = list(rows)
        if summary and summary.get("months"):
            all_rows.append([])
            all_rows.append(["月份", "总支出", "总收入", "结余", "预算", "预算使用率"])
            for m in summary["months"]:
                all_rows.append([
                    m.get("month", ""),
                    m.get("total_expense", 0),
                    m.get("total_income", 0),
                    m.get("net", 0),
                    m.get("budget", 0),
                    f"{m.get('budget_pct', 0):.1f}%",
                ])
        return _write_csv(headers, all_rows)

    # ── Excel: Sheet 1 — 收支明细 ──
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "收支明细"

    ws.append([f"收支明细 — {now_et().strftime('%Y-%m-%d %H:%M')}"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws[1][0].font = Font(name="Arial", bold=True, size=14)
    ws.append([])

    _style_header(ws, headers)

    for row in rows:
        ws.append(row)
        row_idx = ws.max_row
        # 金额格式
        ws.cell(row=row_idx, column=4).number_format = _NUM_FMT_CNY
        # 类型颜色: 支出红色, 收入绿色
        type_val = row[1]
        if type_val == "支出":
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).font = _RED_FONT
        elif type_val == "收入":
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).font = _GREEN_FONT
        # 通用边框
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col_idx).border = _THIN_BORDER

    _auto_width(ws)
    ws.freeze_panes = "A4"

    # ── Excel: Sheet 2 — 月度汇总 ──
    if summary and summary.get("months"):
        ws2 = wb.create_sheet("月度汇总")
        ws2.append(["月度财务汇总"])
        ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        ws2[1][0].font = Font(name="Arial", bold=True, size=14)
        ws2.append([])

        summary_headers = ["月份", "总支出", "总收入", "结余", "预算", "预算使用率"]
        _style_header(ws2, summary_headers)

        for m in summary["months"]:
            net = m.get("net", 0)
            budget_pct = m.get("budget_pct", 0)
            ws2.append([
                m.get("month", ""),
                m.get("total_expense", 0),
                m.get("total_income", 0),
                net,
                m.get("budget", 0),
                budget_pct / 100 if budget_pct else 0,
            ])
            r = ws2.max_row
            # 金额格式
            for c in (2, 3, 4, 5):
                ws2.cell(row=r, column=c).number_format = _NUM_FMT_CNY
            # 百分比格式
            ws2.cell(row=r, column=6).number_format = _NUM_FMT_PCT
            # 结余颜色
            _style_pnl_cell(ws2.cell(row=r, column=4), net)
            # 预算使用率颜色: >100% 红色, >80% 黄色
            pct_cell = ws2.cell(row=r, column=6)
            if budget_pct > 100:
                pct_cell.font = _RED_FONT
                pct_cell.fill = _RED_FILL
            elif budget_pct > 80:
                pct_cell.font = Font(color="9C6500")
                pct_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            # 通用边框
            for c in range(1, 7):
                ws2.cell(row=r, column=c).border = _THIN_BORDER

        _auto_width(ws2)
        ws2.freeze_panes = "A4"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── 闲鱼订单导出 ─────────────────────────────────────────────

def export_xianyu_orders(
    orders: List[dict],
    profit_summary: Optional[dict] = None,
    format: str = "xlsx",
) -> io.BytesIO:
    """导出闲鱼订单 — 订单明细表 + 利润汇总sheet

    Parameters
    ----------
    orders : list[dict]
        每条订单: date, item_name, buyer, status, amount, cost, commission_rate
    profit_summary : dict, optional
        利润汇总: total_orders, total_revenue, total_cost, total_commission, net_profit
    format : str
        "xlsx" (默认) 或 "csv"
    """
    headers = ["日期", "商品", "买家", "状态", "售价", "成本", "佣金", "利润"]

    rows = []
    for o in orders:
        amount = o.get("amount", 0) or 0
        cost = o.get("cost", 0) or 0
        commission_rate = o.get("commission_rate", 0.06) or 0.06
        commission = round(amount * commission_rate, 2)
        profit = round(amount * (1 - commission_rate) - cost, 2)
        rows.append([
            o.get("date", ""),
            o.get("item_name", ""),
            o.get("buyer", ""),
            o.get("status", ""),
            amount,
            cost,
            commission,
            profit,
        ])

    if format == "csv" or not HAS_OPENPYXL:
        all_rows = list(rows)
        if profit_summary:
            all_rows.append([])
            all_rows.append(["总订单", "总营收", "总成本", "总佣金", "净利润"])
            all_rows.append([
                profit_summary.get("total_orders", 0),
                profit_summary.get("total_revenue", 0),
                profit_summary.get("total_cost", 0),
                profit_summary.get("total_commission", 0),
                profit_summary.get("net_profit", 0),
            ])
        return _write_csv(headers, all_rows)

    # ── Excel: Sheet 1 — 订单明细 ──
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "订单明细"

    ws.append([f"闲鱼订单明细 — {now_et().strftime('%Y-%m-%d %H:%M')}"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws[1][0].font = Font(name="Arial", bold=True, size=14)
    ws.append([])

    _style_header(ws, headers)

    for row in rows:
        ws.append(row)
        row_idx = ws.max_row
        # 金额列格式 (售价/成本/佣金/利润)
        for c in (5, 6, 7, 8):
            ws.cell(row=row_idx, column=c).number_format = _NUM_FMT_CNY
        # 利润颜色编码
        profit_val = row[7] or 0
        _style_pnl_cell(ws.cell(row=row_idx, column=8), profit_val)
        # 通用边框
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col_idx).border = _THIN_BORDER

    _auto_width(ws)
    ws.freeze_panes = "A4"

    # ── Excel: Sheet 2 — 利润汇总 ──
    if profit_summary:
        ws2 = wb.create_sheet("利润汇总")
        ws2.append(["闲鱼利润汇总"])
        ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
        ws2[1][0].font = Font(name="Arial", bold=True, size=14)
        ws2.append([])

        # 键值对布局
        summary_items = [
            ("总订单数", profit_summary.get("total_orders", 0), False),
            ("总营收", profit_summary.get("total_revenue", 0), True),
            ("总成本", profit_summary.get("total_cost", 0), True),
            ("总佣金", profit_summary.get("total_commission", 0), True),
            ("净利润", profit_summary.get("net_profit", 0), True),
        ]
        for label, value, is_money in summary_items:
            ws2.append([label, value])
            r = ws2.max_row
            ws2.cell(row=r, column=1).font = Font(bold=True)
            ws2.cell(row=r, column=1).border = _THIN_BORDER
            ws2.cell(row=r, column=2).border = _THIN_BORDER
            if is_money:
                ws2.cell(row=r, column=2).number_format = _NUM_FMT_CNY
            # 净利润颜色
            if label == "净利润":
                _style_pnl_cell(ws2.cell(row=r, column=2), value)

        _auto_width(ws2)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_watchlist(items: List[dict], format: str = "xlsx") -> io.BytesIO:
    """导出自选股列表

    Parameters
    ----------
    items : list[dict]
        每项包含: symbol, reason(可选), added_by(可选), added_at(可选)
    """
    headers = ["代码", "添加原因", "添加人", "添加时间"]

    rows = []
    for item in items:
        rows.append([
            item.get("symbol", ""),
            item.get("reason", ""),
            item.get("added_by", ""),
            item.get("added_at", "")[:19] if item.get("added_at") else "",
        ])

    if format == "csv" or not HAS_OPENPYXL:
        return _write_csv(headers, rows)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "自选股"

    ws.append([f"自选股列表 — {now_et().strftime('%Y-%m-%d %H:%M')}"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws[1][0].font = Font(name="Arial", bold=True, size=14)
    ws.append([])

    _style_header(ws, headers)

    for row in rows:
        ws.append(row)
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=ws.max_row, column=col_idx).border = _THIN_BORDER

    _auto_width(ws)
    ws.freeze_panes = "A4"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_portfolio(
    positions: List[dict],
    summary: Optional[dict] = None,
    format: str = "xlsx",
) -> io.BytesIO:
    """导出投资组合

    Parameters
    ----------
    positions : list[dict]
        每个持仓: symbol, quantity, avg_cost, market_value, pnl_pct
    summary : dict, optional
        现金 / 总资产等汇总信息
    """
    headers = ["代码", "持仓数量", "平均成本", "市值", "盈亏%"]

    rows = []
    for p in positions:
        rows.append([
            p.get("symbol", ""),
            p.get("quantity", 0),
            p.get("avg_cost", 0),
            p.get("market_value", 0),
            p.get("pnl_pct", 0),
        ])

    if format == "csv" or not HAS_OPENPYXL:
        # 在 CSV 末尾追加汇总
        if summary:
            rows.append([])
            rows.append(["现金", summary.get("cash", 0)])
            rows.append(["总资产", summary.get("total_value", 0)])
        return _write_csv(headers, rows)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "投资组合"

    ws.append([f"投资组合 — {now_et().strftime('%Y-%m-%d %H:%M')}"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws[1][0].font = Font(name="Arial", bold=True, size=14)
    ws.append([])

    _style_header(ws, headers)

    for row in rows:
        ws.append(row)
        row_idx = ws.max_row
        ws.cell(row=row_idx, column=3).number_format = _NUM_FMT_USD
        ws.cell(row=row_idx, column=4).number_format = _NUM_FMT_USD
        # 盈亏百分比颜色
        pnl_cell = ws.cell(row=row_idx, column=5)
        pnl_cell.number_format = '0.00"%"'
        _style_pnl_cell(pnl_cell, row[4] or 0)
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col_idx).border = _THIN_BORDER

    # 汇总区域
    if summary:
        ws.append([])
        summary_start = ws.max_row + 1
        ws.append(["现金", "", "", summary.get("cash", 0)])
        ws.cell(row=ws.max_row, column=4).number_format = _NUM_FMT_USD
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        ws.append(["总资产", "", "", summary.get("total_value", 0)])
        ws.cell(row=ws.max_row, column=4).number_format = _NUM_FMT_USD
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)

    _auto_width(ws)
    ws.freeze_panes = "A4"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
